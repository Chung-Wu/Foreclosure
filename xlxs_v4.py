import os
import sys
import openpyxl as op
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import copy

exit_flag = 0

def ws_copy(ws_target, ws_source):
    ws_target.column_dimensions['A'].width = 19.22
    ws_target.column_dimensions['B'].width = 29.22
    ws_target.column_dimensions['C'].width = 29.22
    ws_target.column_dimensions['D'].width = 15.22
    ws_target.column_dimensions['E'].width = 49.22
    ws_target.column_dimensions['F'].width = 19.22
    ws_target.column_dimensions['G'].width = 9.22
    ws_target.column_dimensions['H'].width = 9.22
    ws_target.column_dimensions['I'].width = 9.22
    ws_target.column_dimensions['J'].width = 19.22
    ws_target.column_dimensions['K'].width = 9.22

    for i in range(1, ws_source.max_row+1):
        rs = ws_source.row_dimensions[i].height
        if rs != None:
            ws_target.row_dimensions[i].height = rs

        for j in range(1, ws_source.max_column+1):
            

            target_cell = ws_target.cell(i, j)
            source_cell = ws_source.cell(i, j)
            target_cell.value = source_cell.value
            if(ws_source.cell(row=i, column=j).has_style):
                target_cell._style = copy.copy(source_cell._style)
                target_cell.font = copy.copy(source_cell.font)
                target_cell.border = copy.copy(source_cell.border)
                target_cell.fill = copy.copy(source_cell.fill)
                target_cell.number_format = copy.copy(source_cell.number_format)
                target_cell.protection = copy.copy(source_cell.protection)
                target_cell.alignment = copy.copy(source_cell.alignment)

def one_row_copy(target_row_num, source_row_num, ws_target, ws_source):

    for j in range(1, ws_source.max_column+1):
        target_cell = ws_target.cell(target_row_num, j)
        source_cell = ws_source.cell(source_row_num, j)
        target_cell.value = source_cell.value
        if(ws_source.cell(row=target_row_num, column=j).has_style):
            target_cell._style = copy.copy(source_cell._style)
            target_cell.font = copy.copy(source_cell.font)
            target_cell.border = copy.copy(source_cell.border)
            target_cell.fill = copy.copy(source_cell.fill)
            target_cell.number_format = copy.copy(source_cell.number_format)
            target_cell.protection = copy.copy(source_cell.protection)
            target_cell.alignment = copy.copy(source_cell.alignment)

def remove_repetition(target):
    ws_trace = target.worksheets[1]
    ws_not_trace = target.worksheets[2]
    ws_target = target.worksheets[0]

    print("-------------------------------------------------")
    print("處理 [不追蹤] 清單:")
    print("-------------------------------------------------")
    
    i = 1
    while(True):
        row_data_not_trace = ws_not_trace[i]
    
        if(row_data_not_trace[0].value == None):
            print("第", i, "筆:", row_data_not_trace[0].value)
            break
        elif(row_data_not_trace[0].value == '法院名稱'):
            print("第", i, "筆:", row_data_not_trace[0].value)
            i += 1
            continue
        else:
            print("第", i, "筆:", row_data_not_trace[0].value, row_data_not_trace[4].value.split('\n')[0])
            j = 1
            while(True):
                row_data_target = ws_target[j]
                if(row_data_target[0].value == None):
                    break
                if((row_data_not_trace[0].value == row_data_target[0].value) and (row_data_not_trace[1].value == row_data_target[1].value)):
                    print("target中第", j, "筆資料命中:", row_data_target[0].value, row_data_target[1].value)
                    print("移除第", j, "筆資料")
                    ws_target.delete_rows(j)
                else:
                    j += 1

            i += 1

            
    orange_fill = PatternFill(fill_type='solid', fgColor="FFC125")
    

    print("-------------------------------------------------")
    print("處理 [追蹤] 清單:")
    print("-------------------------------------------------")
    # i = 1

    # while(True):
    #     row_data_trace = ws_trace[i]
        
    #     if(row_data_trace[0].value == None):
    #         print("第", i, "筆:", row_data_trace[0].value)
    #         break
    #     elif(row_data_trace[0].value == '法院名稱'):
    #         print("第", i, "筆:", row_data_trace[0].value)
    #         i += 1
    #         continue
    #     else:

    #         j = 1
    #         while(True):
    #             row_data_target = ws_target[j]
    #             if(row_data_target[0].value == None):
    #                 break
    #             if((row_data_trace[0].value == row_data_target[0].value) and (row_data_trace[1].value == row_data_target[1].value)):
    #                 if(row_data_trace[4].value.split('\n')[0] == row_data_target[4].value.split('\n')[0]):
    #                     for k in range(1, ws_target.max_column+1):
    #                         ws_target.cell(row=j, column=k).fill = orange_fill
    #                     # ws_target.delete_rows(j)
    #             j += 1
    #         i += 1


    
    i = 1

    while(True):
        row_data_trace = ws_trace[i]
        
        if(row_data_trace[0].value == None):
            print("第", i, "筆:", row_data_trace[0].value)
            break
        elif(row_data_trace[0].value == '法院名稱'):
            print("第", i, "筆:", row_data_trace[0].value)
            i += 1
            continue
        else:
            print("第", i, "筆:", row_data_trace[0].value, row_data_trace[4].value.split('\n')[0])
            j = 1
            while(True):
                row_data_target = ws_target[j]
                if(row_data_target[0].value == None):
                    break
                if((row_data_trace[0].value == row_data_target[0].value) and (row_data_trace[1].value == row_data_target[1].value)):

                    if(row_data_trace[4].value.split('\n')[0] == row_data_target[4].value.split('\n')[0]):
                        one_row_copy(i, j, ws_trace, ws_target)
                        ws_target.delete_rows(j)
                    # else:
                    #     print("target中第", j, "筆資料 案號 相同:", row_data_target[0].value, row_data_target[1].value)
                    #     print("移除第", j, "筆資料")
                    #     ws_target.delete_rows(j)
                j += 1
            i += 1

def check_excel_is_open():
    temp_file = './~$target.xlsx'
    if(os.path.exists(temp_file)):
        print("target.xlsx檔案開啟中 關閉才能順利執行程式 ")
        print('輸入任意鍵結束程式')
        input()
        sys.exit()
    else:
        print("target.xlsx檔案沒有開啟 順利執行程式")
            
    

if(os.path.exists("./target.xlsx") == True):
    print("[參考檔案] 存在")
else:
    print("=====================")
    print("[參考檔案] 不存在 記得從網站上下載")
    print("=====================")

    print('輸入任意鍵繼續...')
    sys.stdin.read(1)
    sys.exit()

if(os.path.exists("./trace.xlsx") == True):
    print("[追蹤] 存在")
else:
    exit_flag = 1
    print("=====================")
    print("新增 [追蹤] 檔案")
    print("=====================")
    wb = op.Workbook()
    actsheet = wb.active
    actsheet.title = '追蹤'
    actsheet.column_dimensions['A'].width = 19.22
    actsheet.column_dimensions['B'].width = 29.22
    actsheet.column_dimensions['C'].width = 29.22
    actsheet.column_dimensions['D'].width = 15.22
    actsheet.column_dimensions['E'].width = 49.22
    actsheet.column_dimensions['F'].width = 19.22
    actsheet.column_dimensions['G'].width = 9.22
    actsheet.column_dimensions['H'].width = 9.22
    actsheet.column_dimensions['I'].width = 9.22
    actsheet.column_dimensions['J'].width = 19.22
    actsheet.column_dimensions['K'].width = 9.22

    actsheet["A1"].value = "法院名稱"
    actsheet["B1"].value = "字號股別"
    actsheet["C1"].value = "拍賣日期/拍賣次數"
    actsheet["D1"].value = "縣市"
    actsheet["E1"].value = "土地坐落/面積"
    actsheet["F1"].value = "拍賣總底價(元)"
    actsheet["G1"].value = "點交"
    actsheet["H1"].value = "空地"
    actsheet["I1"].value = "標別"
    actsheet["J1"].value = "備註"
    actsheet["K1"].value = "採通訊投標"
    


    wb.save('trace.xlsx')


if(os.path.exists("./not_trace.xlsx") == True):
    print("[不追蹤] 存在 ")
else:
    exit_flag = 1
    print("=====================")
    print("新增 [不追蹤] 檔案")
    print("=====================")
    wb = op.Workbook()
    actsheet = wb.active
    actsheet.title = '不追蹤'
    actsheet.column_dimensions['A'].width = 19.22
    actsheet.column_dimensions['B'].width = 29.22
    actsheet.column_dimensions['C'].width = 29.22
    actsheet.column_dimensions['D'].width = 15.22
    actsheet.column_dimensions['E'].width = 49.22
    actsheet.column_dimensions['F'].width = 19.22
    actsheet.column_dimensions['G'].width = 9.22
    actsheet.column_dimensions['H'].width = 9.22
    actsheet.column_dimensions['I'].width = 9.22
    actsheet.column_dimensions['J'].width = 19.22
    actsheet.column_dimensions['K'].width = 9.22

    actsheet["A1"].value = "法院名稱"
    actsheet["B1"].value = "字號股別"
    actsheet["C1"].value = "拍賣日期/拍賣次數"
    actsheet["D1"].value = "縣市"
    actsheet["E1"].value = "土地坐落/面積"
    actsheet["F1"].value = "拍賣總底價(元)"
    actsheet["G1"].value = "點交"
    actsheet["H1"].value = "空地"
    actsheet["I1"].value = "標別"
    actsheet["J1"].value = "備註"
    actsheet["K1"].value = "採通訊投標"


    wb.save('not_trace.xlsx')





if(exit_flag == 1):
    print('這次執行後，產生一些必要的檔案，再執行一次程式即可順利執，輸入任意鍵繼續...')
    print("記得打開 trace.xlsx 及 not_trace.xlsx 設定自己希望的格式")
    input('輸入任意鍵繼續...')

    sys.exit()
else:
    check_excel_is_open()
    # open excel files
    wb_target = load_workbook('target.xlsx')
    wb_trace = load_workbook('trace.xlsx')
    wb_not_trace = load_workbook('not_trace.xlsx')

    #create the trace worksheet for target.xlsx
    if not('追蹤' in wb_target.sheetnames):
        ws_trace = wb_target.create_sheet('追蹤')
    else:
        ws_trace = wb_target.worksheets[1]
    if not('不追蹤' in wb_target.sheetnames):    
        ws_not_trace = wb_target.create_sheet('不追蹤')
    else:
        ws_not_trace = wb_target.worksheets[2]

    # copy the worksheet of trace.xlsx to target.xlsx
    source_trace = wb_trace.active              # get the worksheet of trace
    ws_copy(ws_trace, source_trace)
    

    # copy the worksheet of not_trace.xlsx to target.xlsx
    source_not_trace = wb_not_trace.active
    ws_copy(ws_not_trace, source_not_trace)


    remove_repetition(wb_target)
    wb_target.save('target.xlsx')

    print('程式執行完畢 輸入任意鍵繼續')
    input()

