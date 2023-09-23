import os
import sys
import openpyxl as op
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import copy


def ws_copy(ws_target, ws_source):
    for i in range(1, ws_source.max_row+1):
        for j in range(1, ws_source.max_column+1):
            target_cell = ws_target.cell(i, j)
            source_cell = ws_source.cell(i, j)
            target_cell.value = source_cell.value
            if(ws_source.cell(row=i, column=j).has_style):
                target_cell._style = copy.copy(source_cell._style)
                target_cell.font = copy.copy(source_cell.font)
                target_cell.border = copy.copy(source_cell.border)
                target_cell.fill = copy.copy(source_cell.fill)
                target_cell.number_format = copy.copy(source_cell.number_format)
                target_cell.protection = copy.copy(source_cell.protection)
                target_cell.alignment = copy.copy(source_cell.alignment)


def save_files(target):
    ws_trace = target.worksheets[1]
    ws_not_trace = target.worksheets[2]
    wb_trace = load_workbook('trace.xlsx')
    wb_not_trace = load_workbook('not_trace.xlsx')
    
    print("-------------------------------------------------")
    print("複製 [追蹤] 清單:")
    print("-------------------------------------------------")
    ws_copy(wb_trace.active, ws_trace)
    wb_trace.active.title = "追蹤"
    wb_trace.save('trace.xlsx')

    print("-------------------------------------------------")
    print("複製 [不追蹤] 清單:")
    print("-------------------------------------------------")
    ws_copy(wb_not_trace.active, ws_not_trace)
    wb_not_trace.active.title = "不追蹤"
    wb_not_trace.save('not_trace.xlsx')

    


    

def delete_files():
    wb_trace = load_workbook('trace.xlsx')
    # remove all data of wb_trace
    while(True):
        if(wb_trace.active[1][0].value == None):
            break
        else:
            wb_trace.active.delete_rows(1)
    wb_trace.save('trace.xlsx')

    wb_not_trace = load_workbook('not_trace.xlsx')
    # remove all data of wb_not_trace
    while(True):
        if(wb_not_trace.active[1][0].value == None):
            break
        else:
            wb_not_trace.active.delete_rows(1)
    wb_trace.save('not_trace.xlsx')

def check_excel_is_open():
    temp_file = './~$target.xlsx'
    if(os.path.exists(temp_file)):
        print("target.xlsx檔案開啟中 關閉才能順利執行程式 ")
        print('輸入任意鍵結束程式')
        input()
        sys.exit()
    else:
        print("target.xlsx檔案沒有開啟 順利執行程式")

    temp_file = './~$trace.xlsx'
    if(os.path.exists(temp_file)):
        print("trace.xlsx檔案開啟中 關閉才能順利執行程式 ")
        print('輸入任意鍵結束程式')
        input()
        sys.exit()
    else:
        print("trace.xlsx檔案沒有開啟 順利執行程式")

    temp_file = './~$not_trace.xlsx'
    if(os.path.exists(temp_file)):
        print("not_trace.xlsx檔案開啟中 關閉才能順利執行程式 ")
        print('輸入任意鍵結束程式')
        input()
        sys.exit()
    else:
        print("not_trace.xlsx檔案沒有開啟 順利執行程式")


if(os.path.exists("./target.xlsx") == True):
    print("[參考檔案] 存在")
else:
    print("=====================")
    print("[參考檔案] 不存在 記得從網站上下載")
    print("=====================")

    print('輸入任意鍵繼續...')
    sys.stdin.read(1)
    sys.exit()


check_excel_is_open()
wb_target = load_workbook('target.xlsx')

if(len(wb_target.worksheets) != 3):
    print("工作表數量錯誤(應是3)! 程式結束")
    print('程式執行完畢 輸入任意鍵繼續')
    input()
    sys.exit()
else:
    delete_files()
    print("清空 trace.xlsx 及 not_trace.xlsx 檔案")
    save_files(wb_target)

    input("程式執行完畢 輸入任意鍵繼續...")
    sys.exit()

