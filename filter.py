import os
import sys
import openpyxl as op
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import copy


def check_excel_is_open():
    temp_file = './~$target.xlsx'
    if(os.path.exists(temp_file)):
        print("target.xlsx檔案開啟中 關閉才能順利執行程式 ")
        print('輸入任意鍵結束程式')
        input()
        sys.exit()
    else:
        print("target.xlsx檔案沒有開啟 順利執行程式")

    temp_file = './~$trace.xlsx'
    if(os.path.exists(temp_file)):
        print("trace.xlsx檔案開啟中 關閉才能順利執行程式 ")
        print('輸入任意鍵結束程式')
        input()
        sys.exit()
    else:
        print("trace.xlsx檔案沒有開啟 順利執行程式")

    temp_file = './~$not_trace.xlsx'
    if(os.path.exists(temp_file)):
        print("not_trace.xlsx檔案開啟中 關閉才能順利執行程式 ")
        print('輸入任意鍵結束程式')
        input()
        sys.exit()
    else:
        print("not_trace.xlsx檔案沒有開啟 順利執行程式")


if(os.path.exists("./target.xlsx") == True):
    print("[參考檔案] 存在")
else:
    print("=====================")
    print("[參考檔案] 不存在 記得從網站上下載")
    print("=====================")

    print('輸入任意鍵繼續...')
    sys.stdin.read(1)
    sys.exit()



check_excel_is_open()
wb_target = load_workbook('target.xlsx')

ws_target = wb_target.worksheets[0]

i = 1

while(True):
    row_data_target = ws_target[i]

    if(row_data_target[0].value == None):
        break
    elif(row_data_target[0].value == '法院名稱'):
        i += 1
        continue
    else:
        temp_row_data_target = ws_target[i+1]
        if((row_data_target[0].value == temp_row_data_target[0].value) and (row_data_target[1].value == temp_row_data_target[1].value)):
            print("target中第", i, "筆和第:",i+1,"筆資料重複" , row_data_target[0].value, row_data_target[1].value)
            print("移除第", i+1, "筆資料")
            ws_target.delete_rows(i+1)
        else:
            i += 1

wb_target.save('target.xlsx')

print('程式執行完畢 輸入任意鍵繼續')
input()
